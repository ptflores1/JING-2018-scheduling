import openpyxl
from openpyxl.styles import PatternFill
from faker import Faker
import os
fake = Faker()


day_string = {1 : "Viernes", 2 : "Sabado", 3 : "Domingo"}

def by_event(Y, days, events, T_d, instancia):
    wb = openpyxl.Workbook()
    #x[t, e, k, d]
    Y = map(lambda x: (*x[0], x[1].x), Y.items())
    Y = filter(lambda y: y[-1] == 1, Y)
    Y = list(map(lambda x: x[:-1], Y))

    for day in days:
        ws = wb.create_sheet(day_string[day])
        Y_day = filter(lambda d: d[-1] == day, Y)

        # Rellenar celdas de los periodos (filas)
        for i in range(1, T_d[day] + 1):
            ws.cell(i+1, 1, i)

        # Rellenar celdas de los eventos (columnas)
        for event in events:
            ws.cell(1, events.index(event) + 2, event)

        # Rellenar celdas de las canchas (matriz)
        for event in Y_day:
            time, name, field, _ = event
            ws.cell(time + 1, events.index(name) + 2, value=field)
    if not os.path.exists("resultados"):
        os.mkdir("resultados")
    if not os.path.exists("resultados/{}".format(instancia)):
        os.mkdir("resultados/{}".format(instancia))
    del wb["Sheet"]
    wb.save("resultados/{}/calendar_by_event.xlsx".format(instancia))

def by_field(Y, days, fields, T_d, colors, instancia):
    wb = openpyxl.Workbook()
    Y = map(lambda x: (*x[0], x[1].x), Y.items())
    Y = filter(lambda y: y[-1] == 1, Y)
    Y = list(map(lambda x: x[:-1], Y))
    print(Y)

    for day in days:
        ws = wb.create_sheet(day_string[day])
        Y_day = filter(lambda d: d[-1] == day, Y)

        # Rellenar celdas de los periodos (filas)
        for i in range(1, T_d[day] + 1):
            ws.cell(i + 1, 1, i)

        # Rellenar celdas de las canchas (columnas)
        for field in fields:
            ws.cell(1, fields.index(field) + 2, field)

        #Rellenar celdas de los eventos (matriz)
        for event in Y_day:
            time, name, field, _ = event
            c = ws.cell(time + 1, fields.index(field) + 2, value=name)
            c.fill = PatternFill(fill_type="solid", start_color=colors[name].strip("#"))
    if not os.path.exists("resultados"):
        os.mkdir("resultados")
    if not os.path.exists("resultados/{}".format(instancia)):
        os.mkdir("resultados/{}".format(instancia))
    del wb["Sheet"]
    wb.save("resultados/{}/calendar_by_field.xlsx".format(instancia))

def event_colors_by_sport(sports, events_sports):
    s_colors = {s: fake.hex_color() for s in sports}
    e_colors = {e: s_colors[s] for e, s in events_sports.items()}
    return e_colors

