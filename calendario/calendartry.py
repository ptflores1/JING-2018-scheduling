import openpyxl
import pickle

day_string = {1 : "Viernes", 2 : "Sabado", 3 : "Domingo"}

def create_calendar(Y, days, events, T_d):



    wb = openpyxl.Workbook()
    #x[t, e, k, d]
    Y = map(lambda x: (*x[0], x[1].x), Y.items())
    Y = filter(lambda y: y[-1] == 1, Y)
    Y = list(map(lambda x: x[:-1], Y))

    for day in days:
        ws = wb.create_sheet(day_string[day])
        Y_day = filter(lambda d: d[-1] == day, Y)
        for i in range(1, T_d[day] + 1):
            ws.cell(i+1, 1, i)
        for event in events:
            ws.cell(1, events.index(event) + 2, event)
        for event in Y_day:
            time, name, field, _ = event
            ws.cell(time + 1, events.index(name) + 2, value=field)

    wb.save("calendar.xlsx")

